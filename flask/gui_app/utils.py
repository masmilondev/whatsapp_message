from openpyxl import Workbook, load_workbook
from enums import FileType
from whatsapp import WhatsApp

class Utils(WhatsApp):
    def __init__(self, file, fileName, start = 0, end = 0, type=FileType.Customer):
        self.file = file
        self.fileName = fileName
        self.start = start
        self.end = end
        self.type = type
        self.filePath = self.getFilePath()
        super().__init__(self.filePath, self.start, self.end)
    
    def getFilePath(self):
        return self.file.replace("\\","/") +self.fileName+".xlsx"

    def filterCustomer(self, ws):
        filteredContacts = {}
        for i in range(self.start if self.start > 0 else 1, self.end + 1 if self.end > 0 else ws.max_row+1):
            if(ws["D"+str(i)].value == None): continue
            row = [ws["A"+str(i)].value, ws["B"+str(i)].value, ws["D"+str(i)].value]
            filteredContacts[ws["D"+str(i)].value] = row
        return filteredContacts
    
    def filterAbandoned(self, ws):
        completedNumber = []
        for i in range(2, ws.max_row+1):
            if(ws["B" + str(i)].value == "completed"):
                completedNumber.append(ws["E" + str(i)].value)
        print(completedNumber)
        filteredContacts = {}
        for i in range(self.start if self.start > 0 else 1, self.end + 1 if self.end > 0 else ws.max_row+1):
            if(ws["E"+str(i)].value == None): continue

            if(ws["E"+str(i)].value not in completedNumber):
                row = [ws["A"+str(i)].value, ws["C"+str(i)].value, ws["E"+str(i)].value]
                filteredContacts[ws["C"+str(i)].value] = row
        return filteredContacts

    
    def removeDuplicate(self):
        wb = load_workbook(self.filePath)
        ws = wb[FileType.oldSheet.value]
        print("heleor")
        if(self.type == FileType.Customer):
            filteredContacts = self.filterCustomer(ws)
        if(self.type == FileType.abandoned):
            filteredContacts = self.filterAbandoned(ws)

        if(FileType.newSheet.name in wb.sheetnames): wb.remove(wb[FileType.newSheet.name])
        ws = wb.create_sheet(FileType.newSheet.name)

        for j in filteredContacts.values():
            ws.append(j)

        
        wb.save(self.filePath)

    def test(self):
        print("Test okay")