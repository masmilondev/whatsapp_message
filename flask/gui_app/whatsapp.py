from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

import time


from enums import FileType

class WhatsApp:
    def __init__(self, file, start = 0, end = 0):
        self.file = file
        self.start = start
        self.end = end
        self.groupName = "Whatsapp numbers"

    def sendMessage(self, message):
        wb = load_workbook(self.file)
        ws = wb[FileType.newSheet.name]

        fp = webdriver.FirefoxProfile('C:/Users/masmi/AppData/Roaming/Mozilla/Firefox/Profiles/hpoihc25.default-release')
        browser = webdriver.Firefox(fp, executable_path='C:\geckodriver.exe')
        browser.get("https://web.whatsapp.com/")
        time.sleep(7)

        searchBox = browser.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div[3]/div/div[1]/div/label/div/div[2]')
        time.sleep(1.5)
        searchBox.clear()
        # time.sleep(2)
        searchBox.send_keys(self.groupName)
        time.sleep(1.5)

        totalSuccess = 0
        totalError = 0

        if(FileType.errorSheet.name in wb.sheetnames): wb.remove(wb[FileType.errorSheet.name])
        errorSheet = wb.create_sheet(FileType.errorSheet.name)
        if(FileType.successSheet.name in wb.sheetnames): wb.remove(wb[FileType.successSheet.name])
        successSheet = wb.create_sheet(FileType.successSheet.name)

        for i in range(self.start if self.start > 0 else 1, self.end + 1 if self.end > 0 else ws.max_row+1):
            groupMessage = "wa.me/" + str(ws["C"+ str(i)].value)
            try:
                # Find Group name
                groupLink = browser.find_element(By.XPATH, '//span[@title="{}"]'.format(self.groupName))
                time.sleep(1)
                groupLink.click()
                
                # Send number on group
                time.sleep(1)
                messageBox = browser.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]')
                messageBox.send_keys(groupMessage)
                time.sleep(1.5)
                WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button'))).click()
                time.sleep(1.5)

            except NoSuchElementException as se:
                errorSheet.append([ws["A"+ str(i)].value,ws["B"+ str(i)].value,ws["C"+ str(i)].value, "Error"])
                totalError = totalError + 1
                print(groupMessage)
                wb.save(self.file)
                continue
            except ElementClickInterceptedException as se:
                errorSheet.append([ws["A"+ str(i)].value,ws["B"+ str(i)].value,ws["C"+ str(i)].value, "Error"])
                totalError = totalError + 1
                print(groupMessage)
                wb.save(self.file)
                continue
            except Exception as se:
                errorSheet.append([ws["A"+ str(i)].value,ws["B"+ str(i)].value,ws["C"+ str(i)].value, "Error"])
                totalError = totalError + 1
                print(groupMessage)
                wb.save(self.file)
                continue

            print("No exception")
            
            try:
                userName =  browser.find_element(By.XPATH, '//a[@title="{}"]'.format("http://"+ groupMessage))
                userName.click()
                time.sleep(2.5)
                
                messageBox = browser.find_element(By.XPATH, '/html/body/div[1]/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]')
                
                for y in message.splitlines():
                    z = str(y.replace("[customer_name]", ws["B"+ str(i)].value))
                    messageBox.send_keys(z)
                    messageBox.send_keys(Keys.SHIFT + Keys.ENTER)
                    
                WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div/span[2]/div/div[2]/div[2]/button'))).click()
                time.sleep(1)
                successSheet.append([ws["A"+ str(i)].value,ws["B"+ str(i)].value,ws["C"+ str(i)].value, "Success"])
                totalSuccess = totalSuccess + 1
                print(groupMessage)
                print(totalSuccess)
                print(totalError)
                wb.save(self.file)

            except Exception as e:
                time.sleep(2)
                messageBox.clear()
                WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div[2]/div'))).click()
                time.sleep(1)
                print("Error from second try")
                errorSheet.append([ws["A"+ str(i)].value,ws["B"+ str(i)].value,ws["C"+ str(i)].value, "Error"])
                totalError = totalError + 1
                print(groupMessage)
                wb.save(self.file)
                continue

        wb.save(self.file)
        print("Info", """Task completed.
        Total Success: """+str(totalSuccess)+"""
        Total Error: """+str(totalError)+"""
        Please check success and error worksheet""")